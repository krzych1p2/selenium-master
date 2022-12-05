from gpuinfo import GPUInfo
from psutil import virtual_memory
from os import environ
from math import ceil
from psutil import net_if_stats,disk_partitions
from wmi import WMI
from time import strftime,localtime
from openpyxl import Workbook
from openpyxl import load_workbook 
from ssd import is_ssd


# filepath = "dane.txt"  # ścieżka relative
# f = open(filepath, "r")  # otwarcie pliku

# lokalizacja = f.read() 
# f.close()
lokalizacja='podzespoly.xlsx'

def load_to_file():
    try:
        wb = load_workbook(lokalizacja) 
    except:
        wb = Workbook()
    sheet = wb.active
    sheet.append((czas,nazwa_komputer,grupa,uzytkownik,board,procesor,ile_ram,grafika,karta_sieciowa,dysk_ssd))
    wb.save(lokalizacja)
def wyswietl():
    print('1:Time',czas)
    print("2:Computer_name ",nazwa_komputer)
    print('3:Grup =',grupa)
    print('4:User =',uzytkownik)
    print('5:Motherboard =',board )
    print('6:CPU: {0}'.format(procesor))
    print('7:RAM: {0}GB'.format(ile_ram))
    print('8:Graphics Card: {0}'.format(grafika))
    print('9:Network card= {0}GB'.format(karta_sieciowa))
    print("10:SSD data -Disck ", dysk_ssd)

computer = WMI()

#computer_info = computer.Win32_ComputerSystem()[0]

os_info = computer.Win32_OperatingSystem()[0]
czas=strftime("%Y-%m-%d %H:%M:%S", localtime())
nazwa_komputer=environ['COMPUTERNAME']

system_ram = float(os_info.TotalVisibleMemorySize) / 1048576  # KB to GB
ile_ram=ceil(system_ram)

proc_info = computer.Win32_Processor()[0]
procesor = proc_info.Name

gpu_info = computer.Win32_VideoController()[0]
grafika =gpu_info.Name

stats = net_if_stats()
for isup in stats:
    st = stats[isup]
    print(st)
st = stats["Ethernet"]
karta_sieciowa = ceil(st.speed/1024)
#print(stats)
# siec= computer.Win32_NetworkAdapter()[0]
# siec_gotowa=siec.NetConnectionID
# print(siec_gotowa)


#print(plyta_glowna)
#os_name = os_info.Name.encode('utf-8').split(b'|')[0]
#os_version = ' '.join([os_info.Version, os_info.BuildNumber])
#print('OS Name: {0}'.format(os_name))
#print('OS Version: {0}'.format(os_version))
#print('RAM: {0} GB'.format(system_ram))
grupa=input("Do jakiej grupy należy - ")
uzytkownik=input('Uzytkownik komputera - ')

dyski =disk_partitions(all=False)
dysk_ssd=0
## Sprawdzanie czy ssd
for i in range(len(dyski)):
  sciezka=(dyski[i].device)+'/'
  dysk_czy_ssd=(is_ssd(sciezka))
  if dysk_czy_ssd == True:
    dysk_ssd+=1
if dysk_ssd <=1:
    dysk_ssd=0    


motherboard=computer.Win32_BaseBoard()[0]
board=motherboard.Manufacturer + motherboard.Product



load_to_file()
wyswietl()
input("Press Enter to continue...")



